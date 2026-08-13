#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

TABLE_BUCKETS = ("S1-1Big", "S1-1Small", "S1-2", "S2", "S3")
TEST_FIELDS = ("owner_id", "bucket", "group_name", "workplace", "spv", "username", "scrm_1", "scrm_2")


def s(value):
    return str(value or "").strip()


def key(value):
    return s(value).lower()


def table_bucket(bucket):
    bucket = s(bucket)
    if bucket.startswith("S1-2"):
        return "S1-2"
    if bucket.startswith("S2"):
        return "S2"
    if bucket.startswith("S3"):
        return "S3"
    if bucket in ("S1-1Big", "S1-1Small"):
        return bucket
    return None


def read_csv(path):
    with Path(path).open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def csv_fields(path):
    with Path(path).open(newline="", encoding="utf-8-sig") as handle:
        return tuple(csv.DictReader(handle).fieldnames or ())


def source_username(row, source):
    return s(row.get("owner_username") if source == "full" else row.get("username"))


def aggregate(path):
    counts = Counter()
    ignored = Counter()
    total = mapped = 0
    for row in read_csv(path):
        total += 1
        workplace = s(row.get("workplace"))
        raw_bucket = s(row.get("bucket"))
        bucket = table_bucket(raw_bucket)
        if workplace and bucket:
            counts[(workplace, bucket)] += 1
            mapped += 1
        else:
            ignored[(workplace or "<blank>", raw_bucket or "<blank>")] += 1
    return counts, ignored, total, mapped


def cmd_aggregate(args):
    counts, ignored, total, mapped = aggregate(args.csv)
    print(f"total_rows={total}")
    print(f"mapped_rows={mapped}")
    print(f"ignored_rows={total - mapped}")
    for workplace, bucket in sorted(counts):
        print(f"{workplace}\t{bucket}\t{counts[(workplace, bucket)]}")
    if ignored:
        print("ignored:")
        for workplace, bucket in sorted(ignored):
            print(f"{workplace}\t{bucket}\t{ignored[(workplace, bucket)]}")
    return 0


def load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit("openpyxl is required. Install with: pip install --user openpyxl") from exc
    return load_workbook


def compact(value):
    return s(value).replace(" ", "")


def find_title(ws, month, marker):
    month_text = f"{month}月" if month else ""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            text = compact(ws.cell(row, col).value)
            if marker in text and (not month_text or month_text in text):
                return row, col
    raise SystemExit(f"Cannot find table title for month={month or '*'} marker={marker}")


def detect_area(ws, title_row, title_col):
    header_row = title_row + 1
    bucket_cols = {}
    total_col = None
    for col in range(title_col, ws.max_column + 1):
        text = s(ws.cell(header_row, col).value)
        if text == "总计" and bucket_cols:
            total_col = col
            break
        if text in TABLE_BUCKETS:
            bucket_cols[text] = col
    if not total_col or not bucket_cols:
        raise SystemExit(f"Cannot detect table near row={title_row} col={title_col}")
    workplace_col = min(bucket_cols.values()) - 1
    first_row = header_row + 1
    total_row = None
    for row in range(first_row, ws.max_row + 1):
        if s(ws.cell(row, title_col).value) == "总计":
            total_row = row
            break
    if total_row is None:
        raise SystemExit(f"Cannot find total row below row={header_row}")
    return {
        "workplace_col": workplace_col,
        "bucket_cols": bucket_cols,
        "total_col": total_col,
        "first_row": first_row,
        "last_row": total_row - 1,
        "total_row": total_row,
    }


def write_table(ws, area, counts):
    written = {}
    first_col = min(area["bucket_cols"].values())
    last_col = max(area["bucket_cols"].values())
    for row in range(area["first_row"], area["last_row"] + 1):
        workplace = s(ws.cell(row, area["workplace_col"]).value)
        if not workplace:
            continue
        written[workplace] = {}
        for bucket, col in area["bucket_cols"].items():
            value = counts.get((workplace, bucket), 0)
            ws.cell(row, col).value = value if value else None
            written[workplace][bucket] = value
        ws.cell(row, area["total_col"]).value = f"=SUM({ws.cell(row, first_col).coordinate}:{ws.cell(row, last_col).coordinate})"
    for bucket, col in area["bucket_cols"].items():
        has_value = any(ws.cell(row, col).value for row in range(area["first_row"], area["last_row"] + 1))
        ws.cell(area["total_row"], col).value = f"=SUM({ws.cell(area['first_row'], col).coordinate}:{ws.cell(area['last_row'], col).coordinate})" if has_value else None
    ws.cell(area["total_row"], area["total_col"]).value = f"=SUM({ws.cell(area['total_row'], first_col).coordinate}:{ws.cell(area['total_row'], last_col).coordinate})"
    return written


def verify_table(ws, area, counts):
    errors = []
    for row in range(area["first_row"], area["last_row"] + 1):
        workplace = s(ws.cell(row, area["workplace_col"]).value)
        if not workplace:
            continue
        for bucket, col in area["bucket_cols"].items():
            expected = counts.get((workplace, bucket), 0)
            value = ws.cell(row, col).value
            actual = 0 if value in (None, "") else int(value)
            if actual != expected:
                errors.append(f"{workplace}/{bucket}: expected={expected} actual={actual}")
    return errors


def print_matrix(title, matrix):
    print(title)
    for workplace in sorted(matrix):
        parts = [f"{bucket}={value}" for bucket, value in matrix[workplace].items() if value]
        print(f"{workplace}: {', '.join(parts) if parts else '0'}")


def print_ignored(title, ignored):
    if not ignored:
        print(f"{title}=none")
        return
    details = "; ".join(f"{workplace}/{bucket}={count}" for (workplace, bucket), count in sorted(ignored.items()))
    print(f"{title}={details}")


def cmd_update_excel(args):
    load_workbook = load_openpyxl()
    xlsx = Path(args.xlsx)
    test_counts, test_ignored, test_total, test_mapped = aggregate(args.test_csv)
    full_counts, full_ignored, full_total, full_mapped = aggregate(args.full_csv)
    backup = xlsx.with_name(f"{xlsx.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{xlsx.suffix}")
    shutil.copy2(xlsx, backup)
    try:
        wb = load_workbook(xlsx, data_only=False)
        ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[-1]]
        usage_area = detect_area(ws, *find_title(ws, args.month, "waba使用"))
        total_area = detect_area(ws, *find_title(ws, args.month, "总人数"))
        usage_written = write_table(ws, usage_area, test_counts)
        total_written = write_table(ws, total_area, full_counts)
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.calculation.calcMode = "auto"
        except Exception:
            pass
        wb.save(xlsx)
        check = load_workbook(xlsx, data_only=False)
        check_ws = check[ws.title]
        errors = verify_table(check_ws, usage_area, test_counts) + verify_table(check_ws, total_area, full_counts)
        if errors:
            print("VERIFY_FAILED")
            print("\n".join(errors))
            print(f"backup_kept={backup}")
            return 2
        if args.delete_backup:
            os.remove(backup)
            backup_status = f"deleted {backup}"
        else:
            backup_status = f"kept {backup}"
        print(f"saved={xlsx}")
        print(f"backup={backup_status}")
        print(f"test_rows={test_total}, mapped={test_mapped}, ignored={test_total - test_mapped}")
        print(f"full_rows={full_total}, mapped={full_mapped}, ignored={full_total - full_mapped}")
        print_matrix("usage_table", usage_written)
        print_matrix("total_table", total_written)
        print_ignored("test_ignored", test_ignored)
        print_ignored("full_ignored", full_ignored)
        return 0
    except Exception:
        print(f"backup_kept={backup}")
        raise


def filter_rows(rows, bucket=None, workplace=None, group_name=None):
    result = []
    for row in rows:
        if bucket and bucket.lower() != "all" and key(row.get("bucket")) != bucket.lower():
            continue
        if workplace and key(row.get("workplace")) != workplace.lower():
            continue
        if group_name and key(row.get("group_name")) != group_name.lower():
            continue
        if not s(row.get("group_name")) or not s(row.get("bucket")):
            continue
        result.append(row)
    return result


def compare_key(row, same):
    if same == "group":
        return (key(row.get("group_name")),)
    return (key(row.get("group_name")), key(row.get("bucket")))


def user_map(rows, source, same):
    grouped = defaultdict(dict)
    for row in rows:
        grouped[compare_key(row, same)][key(source_username(row, source))] = row
    return grouped


def print_person(row, source):
    name_field = "owner_username" if source == "full" else "username"
    values = [row.get("owner_id", ""), row.get(name_field, ""), row.get("workplace", ""), row.get("group_name", ""), row.get("bucket", ""), row.get("spv", "")]
    print("\t".join(values))


def cmd_diff(args):
    full = read_csv(args.full_csv)
    test = read_csv(args.test_csv)
    full_all = filter_rows(full, args.bucket, None, None)
    full_scoped = filter_rows(full, args.bucket, args.full_workplace, args.group_name)
    test_scoped = filter_rows(test, args.bucket, None, args.group_name)
    full_all_map = user_map(full_all, "full", args.same)
    full_scoped_map = user_map(full_scoped, "full", args.same)
    test_map = user_map(test_scoped, "test", args.same)

    test_missing = []
    for ck, users in test_map.items():
        if ck in full_all_map:
            test_missing.extend(row for name, row in users.items() if name not in full_all_map[ck])
    full_missing = []
    for ck, users in full_scoped_map.items():
        if ck in test_map:
            full_missing.extend(row for name, row in users.items() if name not in test_map[ck])

    print(f"test_has_dataworks_missing={len(test_missing)}")
    for row in sorted(test_missing, key=lambda r: (compare_key(r, args.same), key(source_username(r, "test")))):
        print_person(row, "test")
    print(f"dataworks_has_test_missing={len(full_missing)}")
    for row in sorted(full_missing, key=lambda r: (compare_key(r, args.same), key(source_username(r, "full")))):
        print_person(row, "full")

    test_only = sorted(set(test_map) - set(full_all_map))
    full_only = sorted(set(full_scoped_map) - set(test_map))
    print(f"test_only_groups={len(test_only)}")
    for ck in test_only:
        print("\t".join(ck), len(test_map[ck]))
    print(f"dataworks_only_groups={len(full_only)}")
    for ck in full_only:
        print("\t".join(ck), len(full_scoped_map[ck]))
    return 0


def cmd_append(args):
    full = filter_rows(read_csv(args.full_csv), args.bucket, args.full_workplace, args.group_name)
    test_path = Path(args.test_csv)
    test = read_csv(test_path)
    fields = csv_fields(test_path)
    if fields != TEST_FIELDS:
        raise SystemExit(f"Unexpected test-list columns: {fields}")
    requested = [name.strip() for name in args.usernames.split(",") if name.strip()]
    requested_keys = [key(name) for name in requested]
    existing_users = {key(row.get("username")) for row in test}
    full_by_user = defaultdict(list)
    for row in full:
        full_by_user[key(row.get("owner_username"))].append(row)

    errors = []
    rows_to_add = []
    skipped = []
    for original, user_key in zip(requested, requested_keys):
        if user_key in existing_users:
            skipped.append(original)
            continue
        matches = full_by_user.get(user_key, [])
        if not matches:
            errors.append(f"not found in DataWorks: {original}")
            continue
        if len(matches) > 1:
            choices = "; ".join(f"{m.get('owner_id')}/{m.get('workplace')}/{m.get('group_name')}/{m.get('bucket')}" for m in matches)
            errors.append(f"ambiguous username {original}: {choices}")
            continue
        source = matches[0]
        rows_to_add.append({
            "owner_id": s(source.get("owner_id")),
            "bucket": s(source.get("bucket")),
            "group_name": s(source.get("group_name")),
            "workplace": s(source.get("workplace")),
            "spv": s(source.get("spv")),
            "username": s(source.get("owner_username")),
            "scrm_1": args.scrm1,
            "scrm_2": args.scrm2,
        })
    if errors:
        raise SystemExit("\n".join(errors))

    if rows_to_add:
        needs_newline = False
        if test_path.stat().st_size:
            with test_path.open("rb") as handle:
                handle.seek(-1, os.SEEK_END)
                needs_newline = handle.read(1) not in (b"\n", b"\r")
        with test_path.open("a", newline="", encoding="utf-8") as handle:
            if needs_newline:
                handle.write("\n")
            writer = csv.DictWriter(handle, fieldnames=TEST_FIELDS, lineterminator="\n")
            writer.writerows(rows_to_add)
    print(f"appended={len(rows_to_add)}")
    for row in rows_to_add:
        print(",".join(row[field] for field in TEST_FIELDS))
    print(f"skipped_existing={len(skipped)}")
    for name in skipped:
        print(name)
    return 0


def main():
    parser = argparse.ArgumentParser(description="WABA test-list update utilities")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("aggregate")
    p.add_argument("--csv", required=True)
    p.set_defaults(func=cmd_aggregate)

    p = sub.add_parser("update-excel")
    p.add_argument("--xlsx", required=True)
    p.add_argument("--full-csv", required=True)
    p.add_argument("--test-csv", required=True)
    p.add_argument("--sheet")
    p.add_argument("--month")
    p.add_argument("--delete-backup", action="store_true")
    p.set_defaults(func=cmd_update_excel)

    p = sub.add_parser("diff")
    p.add_argument("--full-csv", required=True)
    p.add_argument("--test-csv", required=True)
    p.add_argument("--bucket", default="all")
    p.add_argument("--full-workplace")
    p.add_argument("--group-name")
    p.add_argument("--same", choices=("group", "group-bucket"), default="group")
    p.set_defaults(func=cmd_diff)

    p = sub.add_parser("append")
    p.add_argument("--full-csv", required=True)
    p.add_argument("--test-csv", required=True)
    p.add_argument("--usernames", required=True)
    p.add_argument("--full-workplace")
    p.add_argument("--bucket", default="all")
    p.add_argument("--group-name")
    p.add_argument("--scrm1", default="")
    p.add_argument("--scrm2", default="waba")
    p.set_defaults(func=cmd_append)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
